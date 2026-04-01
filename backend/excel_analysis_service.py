"""Excel analysis service for QC dashboard."""

from __future__ import annotations

import datetime as dt
import io
import json
import re
import unicodedata
from collections import Counter
from typing import Any


SEMANTIC_ALIASES = {
    "task_code": [
        "task_code", "code", "order_code", "job_code", "id", "task id", "order id",
    ],
    "task_name": [
        "task_name", "title", "name", "task", "job", "order", "ten", "noi_dung",
    ],
    "assignee": [
        "assignee", "staff", "user", "operator", "owner", "pic", "pic edit", "editor",
    ],
    "deadline": [
        "deadline", "due_date", "due", "eta", "end_date", "delivery_date", "han", "ngay giao",
    ],
    "status": [
        "status", "task_status", "state", "progress",
    ],
    "qc_status": [
        "qc_status", "qc", "review_status", "approval", "approve_status",
    ],
    "qc_note": [
        "qc_note", "note", "notes", "comment", "remark", "feedback", "ghi_chu",
    ],
    "qc_owner": [
        "pic_qc", "qc_owner", "reviewer", "qc_by", "qc_pic",
    ],
    "credit": [
        "credit", "credits", "cost", "spend", "usage", "credit_used",
    ],
    "group": [
        "group", "category", "type", "team", "lane",
    ],
}

OPEN_STATUSES = {
    "", "pending", "in_progress", "processing", "queued", "open", "todo", "to_do",
    "waiting", "review", "qc_pending", "rejected",
}
DONE_STATUSES = {
    "done", "completed", "complete", "success", "approved", "closed", "finish", "finished",
    "ok", "qc_approved",
}


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value).strip()


def _safe_float(value: Any) -> float:
    if value in (None, "", 0, "0"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except Exception:
        return 0.0


def _parse_date(value: Any) -> dt.datetime | None:
    if value in (None, "", 0, "0"):
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min)
    text = str(value).strip()
    if not text:
        return None
    for parser in (
        dt.datetime.fromisoformat,
        lambda s: dt.datetime.strptime(s, "%d/%m/%Y"),
        lambda s: dt.datetime.strptime(s, "%d/%m/%Y %H:%M"),
        lambda s: dt.datetime.strptime(s, "%Y-%m-%d"),
        lambda s: dt.datetime.strptime(s, "%Y-%m-%d %H:%M:%S"),
    ):
        try:
            return parser(text)
        except Exception:
            continue
    return None


def _pick_header_row(rows: list[list[Any]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:5]):
        score = sum(1 for cell in row if _safe_text(cell))
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _map_headers(headers: list[str]) -> dict[str, str]:
    mapped: dict[str, str] = {}
    normalized = {_normalize_text(h): h for h in headers if h}
    for semantic, aliases in SEMANTIC_ALIASES.items():
        chosen = None
        for alias in aliases:
            alias_norm = _normalize_text(alias)
            if alias_norm in normalized:
                chosen = normalized[alias_norm]
                break
        if chosen is None:
            for key, original in normalized.items():
                if any(_normalize_text(alias) in key for alias in aliases):
                    chosen = original
                    break
        if chosen:
            mapped[semantic] = chosen
    return mapped


def _status_closed(value: Any) -> bool:
    norm = _normalize_text(value)
    if norm in DONE_STATUSES:
        return True
    if norm in OPEN_STATUSES:
        return False
    return False


def _sheet_to_records(sheet_name: str, rows: list[list[Any]]) -> dict[str, Any]:
    if not rows:
        return {
            "sheet_name": sheet_name,
            "headers": [],
            "semantic_fields": {},
            "records": [],
            "preview_rows": [],
            "row_count": 0,
        }
    header_idx = _pick_header_row(rows)
    headers = [_safe_text(cell) for cell in rows[header_idx]]
    mapped = _map_headers(headers)
    data_rows = rows[header_idx + 1 :]
    preview_rows = [[_safe_text(cell) for cell in row[:12]] for row in rows[: min(len(rows), 8)]]
    records = []
    for row in data_rows:
        if not any(_safe_text(cell) for cell in row):
            continue
        raw = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            raw[header] = row[idx] if idx < len(row) else None
        semantic = {}
        for field, header in mapped.items():
            semantic[field] = raw.get(header)
        records.append({"raw": raw, "semantic": semantic})
    return {
        "sheet_name": sheet_name,
        "headers": headers,
        "semantic_fields": mapped,
        "records": records,
        "preview_rows": preview_rows,
        "row_count": len(records),
    }


def _build_runtime_snapshot(runtime: dict[str, Any]) -> dict[str, Any]:
    daily = runtime.get("daily") or {}
    weekly = runtime.get("weekly") or {}
    monthly = runtime.get("monthly") or {}
    work_tasks = runtime.get("work_tasks") or []
    shifts = runtime.get("shifts") or []
    qc_queue = runtime.get("qc_queue") or []
    users = set()
    for row in work_tasks:
        users.add(str(row.get("user_display") or row.get("user_name") or "").strip())
    for row in shifts:
        users.add(str(row.get("user_display") or row.get("user_name") or "").strip())
    return {
        "daily": {
            "tasks": int(daily.get("count", 0) or 0),
            "videos": int(daily.get("total_videos", 0) or 0),
            "credits": float(daily.get("total_credits", 0) or 0),
        },
        "weekly": {
            "tasks": int(weekly.get("count", 0) or 0),
            "videos": int(weekly.get("total_videos", 0) or 0),
            "credits": float(weekly.get("total_credits", 0) or 0),
        },
        "monthly": {
            "tasks": int(monthly.get("count", 0) or 0),
            "videos": int(monthly.get("total_videos", 0) or 0),
            "credits": float(monthly.get("total_credits", 0) or 0),
        },
        "open_work_tasks": len([row for row in work_tasks if str(row.get("status", "")).lower() == "active"]),
        "shift_reports": len(shifts),
        "qc_pending": len([row for row in qc_queue if str(row.get("status", "")).lower() == "pending"]),
        "runtime_users": sorted([u for u in users if u]),
    }


def _deterministic_analysis(workbook_name: str, active_sheet: dict[str, Any], runtime: dict[str, Any]) -> dict[str, Any]:
    now = dt.datetime.now()
    records = active_sheet.get("records") or []
    mapped = active_sheet.get("semantic_fields") or {}
    runtime_snapshot = _build_runtime_snapshot(runtime)
    overdue_rows = []
    assignee_counter: Counter[str] = Counter()
    overdue_assignee_counter: Counter[str] = Counter()
    workbook_credits = 0.0
    workbook_users = set()
    missing_users = set()

    for row in records:
        semantic = row.get("semantic") or {}
        assignee = _safe_text(semantic.get("assignee"))
        deadline = _parse_date(semantic.get("deadline"))
        status = semantic.get("qc_status") or semantic.get("status") or ""
        credit = _safe_float(semantic.get("credit"))
        task_code = _safe_text(semantic.get("task_code") or semantic.get("task_name"))
        if assignee:
            assignee_counter[assignee] += 1
            workbook_users.add(assignee)
            if assignee not in runtime_snapshot["runtime_users"]:
                missing_users.add(assignee)
        workbook_credits += credit
        if deadline and deadline < now and not _status_closed(status):
            overdue_rows.append({
                "task": task_code or "(unlabeled)",
                "assignee": assignee or "(unassigned)",
                "deadline": deadline.strftime("%Y-%m-%d"),
                "status": _safe_text(status) or "open",
            })
            overdue_assignee_counter[assignee or "(unassigned)"] += 1

    runtime_month = runtime_snapshot["monthly"]
    row_count = int(active_sheet.get("row_count", 0) or 0)
    summary_items = [
        f"Workbook: {workbook_name}",
        f"Active sheet: {active_sheet.get('sheet_name', '-')}",
        f"Mapped fields: {', '.join(sorted(mapped.keys())) if mapped else 'none'}",
        f"Sheet rows analyzed: {row_count}",
        f"Runtime this month: {runtime_month['tasks']} tasks, {runtime_month['videos']} videos, {runtime_month['credits']:.0f} credits",
    ]
    mismatch_items = []
    if row_count:
        diff = runtime_month["tasks"] - row_count
        mismatch_items.append(
            f"Workbook rows vs runtime monthly tasks: {row_count} vs {runtime_month['tasks']} ({diff:+d})"
        )
    else:
        mismatch_items.append("No sheet rows were detected for comparison.")
    if mapped.get("credit"):
        mismatch_items.append(
            f"Workbook planned credit: {workbook_credits:.0f} vs runtime monthly credit: {runtime_month['credits']:.0f}"
        )
    else:
        mismatch_items.append("No mapped credit column found in workbook.")
    if missing_users:
        mismatch_items.append("Workbook assignees not seen in runtime: " + ", ".join(sorted(missing_users)[:8]))
    else:
        mismatch_items.append("All mapped workbook assignees are present in current runtime users.")

    overdue_items = (
        overdue_rows[:8]
        if overdue_rows
        else [{"task": "-", "assignee": "-", "deadline": "-", "status": "No overdue risk detected from mapped deadlines"}]
    )
    staffing_items = []
    if assignee_counter:
        for name, count in assignee_counter.most_common(5):
            overdue_count = overdue_assignee_counter.get(name, 0)
            staffing_items.append(f"{name}: {count} rows assigned, {overdue_count} overdue")
    else:
        staffing_items.append("No mapped assignee column found in workbook.")

    widgets = [
        {
            "id": "workbook-summary",
            "title": "Workbook Summary",
            "metric": str(row_count),
            "tone": "info",
            "body": f"{active_sheet.get('sheet_name', '-')}: {row_count} rows, {len(mapped)} mapped semantic fields.",
        },
        {
            "id": "runtime-mismatch",
            "title": "Runtime Mismatch",
            "metric": f"{runtime_month['tasks'] - row_count:+d}" if row_count else "n/a",
            "tone": "warn" if row_count and runtime_month["tasks"] != row_count else "info",
            "body": mismatch_items[0],
        },
        {
            "id": "overdue-risk",
            "title": "Overdue Risk",
            "metric": str(len(overdue_rows)),
            "tone": "warn" if overdue_rows else "good",
            "body": overdue_items[0]["status"] if overdue_rows else "No overdue items detected.",
        },
        {
            "id": "staffing-insight",
            "title": "Staffing Insight",
            "metric": str(len(assignee_counter)),
            "tone": "info",
            "body": staffing_items[0],
        },
    ]
    return {
        "summary": {
            "title": "Workbook Summary",
            "items": summary_items,
        },
        "mismatch": {
            "title": "Mismatch",
            "items": mismatch_items,
        },
        "overdue_risk": {
            "title": "Overdue Risk",
            "count": len(overdue_rows),
            "items": overdue_items,
        },
        "staffing_insight": {
            "title": "Staffing Insight",
            "items": staffing_items,
        },
        "widgets": widgets,
    }


def _clean_json_text(text: str) -> str:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    match = re.search(r"\{.*\}", cleaned, re.S)
    return match.group(0) if match else cleaned


def _merge_analysis(base: dict[str, Any], refined: dict[str, Any]) -> dict[str, Any]:
    result = dict(base)
    for key in ("summary", "mismatch", "overdue_risk", "staffing_insight"):
        section = dict(base.get(key) or {})
        section.update(refined.get(key) or {})
        result[key] = section
    if isinstance(refined.get("widgets"), list) and refined["widgets"]:
        result["widgets"] = refined["widgets"]
    return result


async def _gpt_refine_analysis(
    workbook_name: str,
    active_sheet: dict[str, Any],
    runtime: dict[str, Any],
    deterministic: dict[str, Any],
    kie_module,
) -> dict[str, Any]:
    runtime_snapshot = _build_runtime_snapshot(runtime)
    preview_rows = active_sheet.get("preview_rows") or []
    prompt_payload = {
        "workbook_name": workbook_name,
        "active_sheet": active_sheet.get("sheet_name"),
        "mapped_fields": active_sheet.get("semantic_fields"),
        "row_count": active_sheet.get("row_count", 0),
        "preview_rows": preview_rows[:8],
        "runtime_snapshot": runtime_snapshot,
        "deterministic_analysis": deterministic,
    }
    system_prompt = (
        "You analyze operational Excel sheets for a production dashboard. "
        "Return strict JSON only. Do not use markdown. "
        "Keys required: summary, mismatch, overdue_risk, staffing_insight, widgets. "
        "summary/mismatch/staffing_insight must contain {title, items}. "
        "overdue_risk must contain {title, count, items}. "
        "widgets must be an array of objects with {id, title, metric, tone, body}. "
        "If evidence is missing, say so explicitly."
    )
    user_prompt = (
        "Analyze this workbook against runtime production data. "
        "Improve the deterministic analysis without inventing facts. "
        "Focus on summary, mismatch, overdue risk, staffing insight, and 4 dashboard widgets.\n\n"
        + json.dumps(prompt_payload, ensure_ascii=False)
    )
    response = await kie_module.chat_completion(
        "gpt-5-4",
        [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        stream=False,
    )
    content = ""
    try:
        content = (
            response.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
        )
    except Exception:
        content = ""
    if not content:
        return deterministic
    try:
        parsed = json.loads(_clean_json_text(content))
    except Exception:
        return deterministic
    return _merge_analysis(deterministic, parsed if isinstance(parsed, dict) else {})


async def analyze_workbook(
    file_bytes: bytes,
    filename: str,
    runtime: dict[str, Any],
    kie_module,
    preferred_sheet: str = "",
) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"Excel parser unavailable: {exc}")

    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row[:24]))
            if len(rows) >= 300:
                break
        sheets.append(_sheet_to_records(sheet_name, rows))

    if not sheets:
        raise RuntimeError("Workbook has no readable sheets")

    active_sheet = None
    preferred_sheet = str(preferred_sheet or "").strip()
    if preferred_sheet:
        for sheet in sheets:
            if sheet["sheet_name"] == preferred_sheet:
                active_sheet = sheet
                break
    if active_sheet is None:
        active_sheet = max(
            sheets,
            key=lambda s: (len(s.get("semantic_fields") or {}), int(s.get("row_count", 0) or 0)),
        )

    deterministic = _deterministic_analysis(filename, active_sheet, runtime)
    refined = deterministic
    try:
        refined = await _gpt_refine_analysis(filename, active_sheet, runtime, deterministic, kie_module)
    except Exception:
        refined = deterministic

    return {
        "file_name": filename,
        "active_sheet": active_sheet.get("sheet_name", ""),
        "semantic_fields": active_sheet.get("semantic_fields", {}),
        "sheets": [
            {
                "sheet_name": sheet.get("sheet_name", ""),
                "row_count": sheet.get("row_count", 0),
                "semantic_fields": sheet.get("semantic_fields", {}),
            }
            for sheet in sheets
        ],
        "preview_rows": active_sheet.get("preview_rows", []),
        **refined,
    }
